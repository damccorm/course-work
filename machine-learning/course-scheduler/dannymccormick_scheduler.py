"""
Name: Daniel McCormick
Architectural:
	This code has 2 basic stages, managed by course_scheduler. In the first step, course_scheduler calls create_satisfying_schedule,
which is responsible for creating a schedule that meets all goal conditions without violating any constraints. It does this with a DFS
backtracking solution in which it chooses a goal condition, assigns it a valid semester, adds all of its prereqs to the goal conditions,
and recursively continues. If at any point a valid schedule can't be found, it tries assigning the course to another semester as it backtracks up.
	This includes an optional heuristic which constrains the search by only allowing it to add courses above a minimum threshold. This starts at 8
and goes down by 1 until a valid schedule is found or it reaches 0.
FUNCTIONALITY:
	In general, this is successful in finding solutions where they exist. It takes very long to output when solutions don't exist though (there may be
an infinite loop somewhere), and the heuristic takes a long time as a result. Thus it is currently turned off.
"""


import re
from collections import namedtuple
from openpyxl import load_workbook
import time
import pprint

def create_course_dict():
    """
    Creates a dictionary containing course info.
    Keys: namedtuple of the form ('program, designation')
    Values: namedtuple of the form('name, prereqs, credits')
            prereqs is a tuple of prereqs where each prereq has the same form as the keys
    """
    wb = load_workbook('newcatalog.xlsx')
    catalog = wb.get_sheet_by_name('catalog')
    Course = namedtuple('Course', 'program, designation')
    CourseInfo = namedtuple('CourseInfo', 'credits, terms, prereqs')
    course_dict = {}
    for row in range(1, catalog.max_row + 1):
        key = Course(get_val(catalog, 'A', row), get_val(catalog, 'B', row))
        prereqs = tuple(tuple(get_split_course(prereq) for prereq in prereqs.split())
                   for prereqs in none_split(get_val(catalog, 'E', row)))
        val = CourseInfo(get_val(catalog, 'C', row), tuple(get_val(catalog, 'D', row).split()), prereqs)
        course_dict[key] = val
    return course_dict


def get_split_course(course):
    """
    Parses a course from programdesignation into the ('program, designation') form.
    e.g. 'CS1101' -> ('CS', '1101')
    """
    return tuple(split_course for course_part in re.findall('((?:[A-Z]+-)?[A-Z]+)(.+)', course)
                 for split_course in course_part)


def none_split(val):
    """Handles calling split on a None value by returning the empty list."""
    return val.split(', ') if val else ()


def get_val(catalog, col, row):
    """Returns the value of a cell."""
    return catalog[col + str(row)].value


def print_dict(dict):
    """Simply prints a dictionary's key and values line by line."""
    for key in dict:
        print(key, dict[key])

"""
All above code is from the Artificial Intelligence TA.
"""

def course_scheduler (course_descriptions, goal_conditions, initial_state):
	"""
	PRE:
		course_descriptions: Dictionary of courses with the following form.
			key - Course(program='<department>', designation='<number>') e.g. Course(program='CS', designation='2201')
			value - CourseInfo(credits='<# credit hours>', terms=<terms list>, prereqs=<prereqs list>) 
				e.g. CourseInfo(credits='3', terms=('Spring', 'Fall'), prereqs=((('CS', '1101'),),))
		goal_conditions: List of post-conditions after the courses have been taken. 
			e.g. [(‘CS’, ‘mathematics’), (‘CS’, ‘core’), (‘MATH’, ’3641’), (‘CS’, ’1151’), (‘MATH’, ‘2410’)]
		initial_state: List of pre-condition courses that have already been credited to the student.
			e.g. [('CS', '1101'), ('SPAN', '1101')]
	POST:
		Returns a set of scheduled courses that satisfy all of the goal conditions.
			Scheduled courses are of the form (Course, scheduled term, number of credits)
			e.g. ((“CS”, “2201”), (“Spring”, “Frosh”), 3)
		Each semester must have 12-18 hours scheduled. If no such satisfying set exists (in 4 years of semesters using
		only fall and spring terms), returns an empty set.
	"""
	include_heuristic = False
	goal_conditions_map = create_goal_tuples(goal_conditions)
	initial_state_map = add_semester_field(initial_state, 0)
	tot_hours_per_semester = {}
	for i in range (1,9):
		tot_hours_per_semester[i] = 0
	min_allowable_semester = 1
	if include_heuristic:
		min_allowable_semester = 8
	schedule = {}
	while min_allowable_semester > 0 and schedule == {}:
		print (min_allowable_semester)
		schedule, tot_hours_per_semester = create_satisfying_schedule(course_descriptions, goal_conditions_map, initial_state_map, tot_hours_per_semester, min_allowable_semester)
		min_allowable_semester -= 1
	if len(schedule) == 0:
		return schedule
	schedule = fill_courseload(course_descriptions, schedule, tot_hours_per_semester)
	print (format_schedule(schedule, course_descriptions))
	return format_schedule(schedule, course_descriptions)

def create_goal_tuples (goal_conditions):
	"""
	PRE:
		goal_conditions: Set of post conditions after all courses have been taken
		semester: The latest semester the goal can be filled
	POST:
		Returns list of tuples mapping post conditions after all courses have been taken to the latest semester they can be fulfilled
		e.g. [((‘CS’, ‘mathematics’), 8), ((‘CS’, ‘core’), 8), ((‘MATH’, ’3641’), 8), ((‘CS’, ’1151’), 8), ((‘MATH’, ‘2410’), 8)]
	""" 
	new_goal_conditions = []
	for goal in goal_conditions:
		new_goal_conditions.append([goal, 8])
	return new_goal_conditions

def format_schedule(schedule, course_descriptions):
	"""
	PRE: 
		completed_courses: Map mapping scheduled courses to the semester they are to be completed. A zero indicates they are pre-conditions
			e.g. {('CS', '1101'): 1, ('SPAN', '1101'): 0, ('CS', '2201'): 2}
		course_descriptions: Dictionary of courses with the following form.
			key - Course(program='<department>', designation='<number>') e.g. Course(program='CS', designation='2201')
			value - CourseInfo(credits='<# credit hours>', terms=<terms list>, prereqs=<prereqs list>) 
				e.g. CourseInfo(credits='3', terms=('Spring', 'Fall'), prereqs=((('CS', '1101'),),))
	POST:
		Schedule map with each course mapped to a tuple with the following configuration (credits, (semester, year), prereqs (empty))
	"""
	formatted_schedule = {}
	for course in schedule:
		semester_number = schedule[course]
		if semester_number > 0:
			semester = "Spring"
			year = "Frosh"
			if semester_number % 2 == 1:
				semester = "Fall"
				semester_number += 1
			semester_number = semester_number / 2
			if semester_number == 2:
				year = "Sophomore"
			elif semester_number == 3:
				year = "Junior"
			elif semester_number == 4:
				year = "Senior"
			formatted_schedule[(course[0], course[1])] = (course_descriptions[course][0], (semester, year), ())
	return formatted_schedule

def create_satisfying_schedule (course_descriptions, goal_conditions, completed_courses, tot_hours_per_semester, min_allowable_semester):
	"""
	PRE:
		course_descriptions: Dictionary of courses with the following form.
			key - Course(program='<department>', designation='<number>') e.g. Course(program='CS', designation='2201')
			value - CourseInfo(credits='<# credit hours>', terms=<terms list>, prereqs=<prereqs list>) 
				e.g. CourseInfo(credits='3', terms=('Spring', 'Fall'), prereqs=((('CS', '1101'),),))
		goal_conditions: List of tuples mapping courses to be taken to the latest semester in which they can be taken given the
			current set of completed courses. Numbers are used to represent the semesters (1="Fall freshman", 2="Spring freshman"...)
			e.g. [((‘CS’, ‘mathematics’): 8), ((‘CS’, ‘core’), 8), ((‘MATH’, ’3641’), 7), ((‘CS’, ’1151’), 7), ((‘MATH’, ‘2410’), 3)]
		completed_courses: Map mapping scheduled courses to the semester they are to be completed. A zero indicates they are pre-conditions
			e.g. {('CS', '1101'): 1, ('SPAN', '1101'): 0, ('CS', '2201'): 2}
		tot_hours_per_semester: Map mapping each semester (1-8) to the total number of hours taken in that semseter.
			e.g. {1: 0, 2: 6, 3: 3, 4: 10, 5: 7, 6: 6, 7: 6, 8: 9}
	POST:
		Returns 2 dictionaries. The first maps courses to the semester that they will be taken. The second maps semesters to the total number
			of hours taken that semester. All prerequisites are satisfied and no semester has more than 18 hours.
	"""
	if len(goal_conditions) == 0:
		return completed_courses, tot_hours_per_semester
	first_goal = goal_conditions[-1]
	del goal_conditions[-1]
	goal_class = first_goal[0]
	goal_class_info = course_descriptions[goal_class]
	semester_assignment = first_goal[1]
	if goal_class in completed_courses:
		if completed_courses[goal_class] <= semester_assignment:
			# No need to reassign it
			return create_satisfying_schedule (course_descriptions, goal_conditions, completed_courses, tot_hours_per_semester, min_allowable_semester)
		else:
			# It was assigned wrong the first time, backtrack
			return {}, {}

	while semester_assignment >= min_allowable_semester:
		print (tot_hours_per_semester[8], goal_class, len(goal_class_info[2]))
		if is_valid_semester_assignment(goal_class_info, completed_courses, tot_hours_per_semester, goal_class, semester_assignment):
			schedule = assign_course_to_semester(goal_class, goal_conditions, dict(completed_courses), dict(tot_hours_per_semester), goal_class_info, semester_assignment, course_descriptions, min_allowable_semester)
			if len(schedule[0]) > 0:
				return schedule
		semester_assignment -= 1
	return {}, {}

def is_valid_semester_assignment(course_info, completed_courses, tot_hours_per_semester, course, semester_assignment):
	"""
	PRE:
		course_info: tuple of information about the course extracted from the course catalog with the following configuration: (credit hours, (semester, year), prereqs)
		completed_courses: Map mapping scheduled courses to the semester they are to be completed. A zero indicates they are pre-conditions
			e.g. {('CS', '1101'): 1, ('SPAN', '1101'): 0, ('CS', '2201'): 2}
		tot_hours_per_semester: : Map mapping each semester (1-8) to the total number of hours taken in that semseter.
			e.g. {1: 0, 2: 6, 3: 3, 4: 10, 5: 7, 6: 6, 7: 6, 8: 9}
		course: The class we're looking at
		semester_assignment: The proposed semester to assign course
	POST:
		Returns whether assigning the course to that semester will violate any existing conditions
	"""
	if int(course_info[0]) + tot_hours_per_semester[semester_assignment] > 18:
		print ("FALSE")
		return False
	if course in completed_courses and completed_courses[course] != semester_assignment:
		print ("FALSE")
		return False
	if len(course_info[2]) > 0 and semester_assignment == 1:
		print ("FALSE")
		return False
	if semester_assignment % 2 == 0:
		return 'Spring' in course_info[1]
	return 'Fall' in course_info[1]

def assign_course_to_semester(course, goal_conditions, completed_courses, tot_hours_per_semester, course_info, semester, course_descriptions, min_allowable_semester):
	"""
	PRE:
		course: The class we're looking at
		goal_conditions: List of tuples mapping courses to be taken to the latest semester in which they can be taken given the
			current set of completed courses. Numbers are used to represent the semesters (1="Fall freshman", 2="Spring freshman"...)
			e.g. [((‘CS’, ‘mathematics’): 8), ((‘CS’, ‘core’), 8), ((‘MATH’, ’3641’), 7), ((‘CS’, ’1151’), 7), ((‘MATH’, ‘2410’), 3)]
		completed_courses: Map mapping scheduled courses to the semester they are to be completed. A zero indicates they are pre-conditions
			e.g. {('CS', '1101'): 1, ('SPAN', '1101'): 0, ('CS', '2201'): 2}
		tot_hours_per_semester: : Map mapping each semester (1-8) to the total number of hours taken in that semseter.
			e.g. {1: 0, 2: 6, 3: 3, 4: 10, 5: 7, 6: 6, 7: 6, 8: 9}
		course_info: tuple of information about the course extracted from the course catalog with the following configuration: (credit hours, (semester, year), prereqs)
		semester: The proposed semester to assign course
		course_descriptions: Dictionary of courses with the following form.
			key - Course(program='<department>', designation='<number>') e.g. Course(program='CS', designation='2201')
			value - CourseInfo(credits='<# credit hours>', terms=<terms list>, prereqs=<prereqs list>) 
				e.g. CourseInfo(credits='3', terms=('Spring', 'Fall'), prereqs=((('CS', '1101'),),))
	POST: 
		Assigns the course to the schedule and continues the scheduling process. Returns {} if no valid schedule exists with the current course assignments, otherwise
		returns the schedule.
	"""
	completed_courses[course] = semester
	tot_hours_per_semester[semester] += int(course_info[0])
	latest_prereq_semester = semester
	if int(course_info[0]) != 0:
		latest_prereq_semester -= 1
	if len(course_info[2]) == 0:
		return create_satisfying_schedule(course_descriptions, list(goal_conditions), dict(completed_courses), dict(tot_hours_per_semester), min_allowable_semester)
	for prereq_set in course_info[2]:

		new_goal_conditions = list(goal_conditions)
		for prereq in prereq_set:
			new_goal_conditions.append((prereq, latest_prereq_semester))
		schedule = add_prereq_set(prereq_set, course_descriptions, list(goal_conditions), dict(completed_courses), dict(tot_hours_per_semester), latest_prereq_semester, min_allowable_semester)
		if len(schedule[0]) > 0:
			return schedule
	return {}, {}

def add_prereq_set (prereq_set, course_descriptions, goal_conditions, completed_courses, tot_hours_per_semester, latest_prereq_semester, min_allowable_semester):
	"""
	PRE:
		prereq_set: The courses to add to the goal conditions in latest_prereq_semester
		course_descriptions: Dictionary of courses with the following form.
			key - Course(program='<department>', designation='<number>') e.g. Course(program='CS', designation='2201')
			value - CourseInfo(credits='<# credit hours>', terms=<terms list>, prereqs=<prereqs list>) 
				e.g. CourseInfo(credits='3', terms=('Spring', 'Fall'), prereqs=((('CS', '1101'),),))
		goal_conditions: List of tuples mapping courses to be taken to the latest semester in which they can be taken given the
			current set of completed courses. Numbers are used to represent the semesters (1="Fall freshman", 2="Spring freshman"...)
			e.g. [((‘CS’, ‘mathematics’): 8), ((‘CS’, ‘core’), 8), ((‘MATH’, ’3641’), 7), ((‘CS’, ’1151’), 7), ((‘MATH’, ‘2410’), 3)]
		completed_courses: Map mapping scheduled courses to the semester they are to be completed. A zero indicates they are pre-conditions
			e.g. {('CS', '1101'): 1, ('SPAN', '1101'): 0, ('CS', '2201'): 2}
		tot_hours_per_semester: : Map mapping each semester (1-8) to the total number of hours taken in that semseter.
			e.g. {1: 0, 2: 6, 3: 3, 4: 10, 5: 7, 6: 6, 7: 6, 8: 9}
		latest_prereq_semester: The latest semester the prereqs can be assigned.
	POST: 
		Adds the prereqs to the goal conditions (if allowable) and continues the scheduling process, returning a valid schedule if one is found.
	"""
	for prereq in prereq_set:
		if prereq in completed_courses:
			if completed_courses[prereq] > latest_prereq_semester:
				return {},{}
		else:
			found = False
			for i in range(len(goal_conditions)):
				if goal_conditions[i][0] == prereq:
					print ("ASDFFFFFFASDFFFFFFF")
					found = True
					goal_conditions[i][1] = min(latest_prereq_semester, goal_conditions[i][1])
			if not found:
					goal_conditions.append([prereq, latest_prereq_semester])
	return create_satisfying_schedule(course_descriptions, goal_conditions, completed_courses, tot_hours_per_semester, min_allowable_semester)

def add_semester_field (goal_conditions, semester):
	"""
	PRE:
		goal_conditions: Set of post conditions after all courses have been taken
		semester: The latest semester the goal can be filled
	POST:
		Returns map of post conditions after all courses have been taken mapped to the latest semester they can be fulfilled
		e.g. {(‘CS’, ‘mathematics’): 8 (‘CS’, ‘core’): 8 (‘MATH’, ’3641’): 8 (‘CS’, ’1151’): 8 (‘MATH’, ‘2410’): 8}
	""" 
	updated_conditions = {}
	for condition in goal_conditions:
		updated_conditions[condition] = semester
	return updated_conditions

def fill_courseload (course_descriptions, schedule, tot_hours_per_semester):
	"""
	PRE:
		course_descriptions: Dictionary of courses with the following form.
			key - Course(program='<department>', designation='<number>') e.g. Course(program='CS', designation='2201')
			value - CourseInfo(credits='<# credit hours>', terms=<terms list>, prereqs=<prereqs list>) 
				e.g. CourseInfo(credits='3', terms=('Spring', 'Fall'), prereqs=((('CS', '1101'),),))
		schedule: Map mapping scheduled courses to the semester they are to be completed. The schedule may be incomplete
		tot_hours_per_semester: Map mapping each semester (1-8) to the total number of hours taken in that semseter.
			e.g. {1: 0, 2: 6, 3: 3, 4: 10, 5: 7, 6: 6, 7: 6, 8: 9}
	POST:
		Returns a valid schedule building off of the input schedule where all semesters have 12-18 hours.
	"""
	i = 1
	while tot_hours_per_semester[i] == 0 or tot_hours_per_semester[i] >= 12:
		i += 1
		if i == 9:
			return schedule
	while i < 9:
		for course in course_descriptions:
			course_info = course_descriptions[course]
			if is_valid_class(course, int(course_info[0]), course_info[1], course_info[2], schedule, tot_hours_per_semester, i):
				schedule[course] = i
				tot_hours_per_semester[i] += int(course_info[0])
				while tot_hours_per_semester == 0 or tot_hours_per_semester[i] >= 12:
					i += 1
					if i == 9:
						return schedule
	return schedule

def is_valid_class (course, credit_hours, terms, prereqs, schedule, tot_hours_per_semester, semester):
	"""
	PRE:
		course: The course to determine the validity of.
		credit_hours: Integer representing the number of hours the course counts for.
		terms: Tuple that may contain 'Fall', 'Spring', and/or 'Summer' representing which terms a course can be taken in.
		prereqs: Tuple of tuples of satisfying prereqs to the course.
		schedule: Dictionary representing the current schedule.
		tot_hours_per_semester: Map mapping each semester (1-8) to the total number of hours taken in that semseter.
		semester: Integer representing the semester to assign the course.
	POST:
		Returns true if class can be assigned to that semester, false if it can't be.
	"""
	if course in schedule:
		return False
	if credit_hours + tot_hours_per_semester[semester] > 18:
		return False
	if semester % 2 == 0:
		if 'Spring' not in terms:
			return False
	elif 'Fall' not in terms:
		return False
	if len(prereqs) == 0:
		return True
	for req_set in prereqs:
		satisfying = True
		for req in prereqs:
			if req not in schedule or schedule[req] >= semester:
				satisfying = False
		if satisfying:
			return True
	return False

def main():
	# TODO: Add heuristic portion
	Course = namedtuple('Course', 'program, designation')
	musl = Course('MUSL', '2110')
	goal_conditions = [Course('CS', 'major')]
	initial_state = [Course('CS', '1101'), Course('JAPN', '1101')]
	plan = course_scheduler(create_course_dict(), goal_conditions, initial_state)
	pp = pprint.PrettyPrinter()

	pp.pprint (plan)

if __name__ == "__main__":
    main()